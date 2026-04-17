"""Combine three sources — More in Common Apr 2026 MRP, Electoral Calculus Jan 2026 MRP,
and 2024 HoC Library results — into one per-constituency CSV with swing and an
empirical per-seat σ derived from cross-pollster disagreement."""

import csv
import math
import re
import unicodedata
import openpyxl

MIC_XLSX = "apr26-mrp-datatables.xlsx"   # More in Common April 2026
EC_XLSX  = "ec_jan2026.xlsx"             # Electoral Calculus January 2026
GE2024   = "ge2024_results.csv"          # HoC Library 2024 results
OUT      = "reform_constituency_predictions.csv"

PARTIES_MIC = ["Conservative", "Labour", "Liberal Democrat", "Reform UK",
               "The Green Party", "Scottish National Party (SNP)",
               "Plaid Cymru", "Other"]

def norm(s):
    if not s: return ""
    # strip diacritics (Ynys Môn -> Ynys Mon)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return re.sub(r'[^a-z0-9]+', '', s.lower())

DIRECTIONS = {"north", "south", "east", "west", "mid", "central", "upper", "lower"}

# Hand-coded aliases for mismatches that don't fit the direction-swap pattern.
MANUAL_ALIASES = {
    # EC strips 'Kingston upon' from the three Hull seats
    "Kingston upon Hull East": ["Hull East"],
    "Kingston upon Hull North and Cottingham": ["Hull North and Cottingham"],
    "Kingston upon Hull West and Haltemprice": ["Hull West and Haltemprice"],
    # EC uses comma-inverted form
    "City of Durham": ["Durham City of", "Durham, City of"],
    "The Wrekin": ["Wrekin The", "Wrekin, The"],
    # EC rearranges directions after the placename even with compound directions
    "Mid and South Pembrokeshire": ["Pembrokeshire Mid and South"],
}

def name_variants(name):
    """Possible normalised keys for a constituency name, handling:
    - EC 'Place Direction' vs MiC 'Direction Place' ordering per clause
    - Parenthetical duplicates ('Caerfyrddin (Carmarthen)')
    - Diacritics ('Ynys Môn' / 'Ynys Mon')
    - Hand-coded aliases for edge cases (Hull seats, Durham, Wrekin)."""
    if not name: return set()
    base = re.sub(r'\s*\([^)]*\)', '', name).strip()
    forms = {base, *MANUAL_ALIASES.get(name, []), *MANUAL_ALIASES.get(base, [])}
    parts = base.split(" and ")
    for i, part in enumerate(parts):
        toks = part.split()
        if toks and toks[0].lower() in DIRECTIONS:
            new = " ".join(toks[1:]) + " " + toks[0]
            np = list(parts); np[i] = new
            forms.add(" and ".join(np))
        if len(toks) >= 2 and toks[-1].lower() in DIRECTIONS:
            new = toks[-1] + " " + " ".join(toks[:-1])
            np = list(parts); np[i] = new
            forms.add(" and ".join(np))
    return {norm(f) for f in forms}

def phi(z):
    return 0.5 * (1.0 + math.erf(z / math.sqrt(2.0)))

def categorise(p):
    if p >= 0.85: return "Safe Reform"
    if p >= 0.60: return "Likely Reform"
    if p >= 0.40: return "Toss-up"
    if p >= 0.15: return "Lean Against"
    return "Safe Against"

# --- 1. Load More in Common ---------------------------------------------
wb = openpyxl.load_workbook(MIC_XLSX, data_only=True)
ws = wb["Results"]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
idx = {p: header.index(p) for p in PARTIES_MIC}
winner_idx = header.index("Winner")
const_idx = header.index("Constituency")

mic = {}
for r in rows[1:]:
    name = r[const_idx]
    if not name: continue
    shares = {p: (r[idx[p]] or 0.0) for p in PARTIES_MIC}
    mic[norm(name)] = {
        "name": name,
        "shares": shares,
        "winner": r[winner_idx],
        "variants": name_variants(name),
    }
print(f"MiC: {len(mic)} constituencies")

def lookup(table, variants):
    for v in variants:
        if v in table:
            return table[v]
    return None

# --- 2. Load Electoral Calculus (Jan 2026) ------------------------------
wb = openpyxl.load_workbook(EC_XLSX, data_only=True)
ws = wb["Seats"]
ec = {}
# data rows start at row 20; columns B..O (index 1..14)
# B=Seat Name, D=CON, E=LAB, F=LIB, G=Reform, H=Green, I=YP, J=SNP/Plaid,
# K=Minor, L=Indep/Other, N=Predicted Winner (with TV), O=Winner 2024
for row in ws.iter_rows(min_row=20, values_only=True):
    name = row[1]
    if not name or not isinstance(name, str) or name.strip() == "":
        continue
    if isinstance(row[3], (int, float)):
        con, lab, lib, reform, green, yp, snp_pc, minor, indep = [
            (row[i] or 0.0) for i in (3, 4, 5, 6, 7, 8, 9, 10, 11)
        ]
    else:
        continue
    predicted = row[13]
    # EC labels parties differently — map to the non-Reform best share
    other_shares = {
        "Conservative": con, "Labour": lab, "Liberal Democrat": lib,
        "Green": green, "YP": yp, "SNP/Plaid": snp_pc, "Minor": minor,
        "Independent/Other": indep,
    }
    rec = {
        "name": name,
        "reform_share": reform,
        "other_shares": other_shares,
        "winner": predicted,
    }
    for v in name_variants(name):
        ec[v] = rec
print(f"EC: {len({id(v) for v in ec.values()})} distinct constituencies "
      f"({len(ec)} name-variants indexed)")

# --- 3. Load 2024 election results --------------------------------------
g24 = {}
with open(GE2024) as f:
    for r in csv.DictReader(f):
        name = r["Constituency name"]
        valid = float(r["Valid votes"] or 0)
        if valid <= 0: continue
        ruk = float(r["RUK"] or 0)
        rec = {
            "name": name,
            "reform_share_2024": ruk / valid,
            "winner_2024": r["First party"],
        }
        for v in name_variants(name):
            g24[v] = rec
print(f"2024: {len({id(v) for v in g24.values()})} distinct constituencies")

# --- 4. Name-matching diagnostics ---------------------------------------
matched_ec  = sum(1 for rec in mic.values() if lookup(ec, rec["variants"]))
matched_g24 = sum(1 for rec in mic.values() if lookup(g24, rec["variants"]))
print(f"\nMiC ∩ EC:   {matched_ec} / {len(mic)}")
print(f"MiC ∩ 2024: {matched_g24} / {len(mic)}")
unmatched = [rec["name"] for rec in mic.values()
             if not lookup(ec, rec["variants"])]
if unmatched:
    print(f"Still unmatched vs EC ({len(unmatched)}): {unmatched[:10]}")

# --- 5. Compute cross-pollster margin disagreement σ ---------------------
diffs = []
for mic_row in mic.values():
    ec_row = lookup(ec, mic_row["variants"])
    if not ec_row: continue
    r_mic = mic_row["shares"]["Reform UK"]
    top_mic = max(v for p, v in mic_row["shares"].items() if p != "Reform UK")
    margin_mic = r_mic - top_mic
    margin_ec = ec_row["reform_share"] - max(ec_row["other_shares"].values())
    diffs.append(abs(margin_mic - margin_ec) * 100)

rms = math.sqrt(sum(d*d for d in diffs) / len(diffs))
# |m1 - m2| has SD ≈ σ_margin * √2, so σ_floor = RMS / √2
SIGMA_FLOOR = rms / math.sqrt(2)
print(f"\nCross-pollster margin-disagreement RMS: {rms:.2f} pp")
print(f"Derived σ_floor: {SIGMA_FLOOR:.2f} pp")

# --- 6. Build output rows -----------------------------------------------
out_rows = []
for k, mic_row in mic.items():
    r_mic = mic_row["shares"]["Reform UK"]
    others_mic = {p: v for p, v in mic_row["shares"].items() if p != "Reform UK"}
    top_mic_party = max(others_mic, key=others_mic.get)
    top_mic_share = others_mic[top_mic_party]
    margin_mic = r_mic - top_mic_share

    ec_row = lookup(ec, mic_row["variants"])
    g24_row = lookup(g24, mic_row["variants"])

    if ec_row:
        r_ec = ec_row["reform_share"]
        margin_ec = r_ec - max(ec_row["other_shares"].values())
        r_mean = (r_mic + r_ec) / 2
        margin_mean = (margin_mic + margin_ec) / 2
        diff_pp = abs(margin_mic - margin_ec) * 100
        sigma_i = math.sqrt(SIGMA_FLOOR ** 2 + (diff_pp / math.sqrt(2)) ** 2)
        winners_agree = (mic_row["winner"] == "Reform UK") == (
            ec_row["winner"] == "Reform")
    else:
        r_ec = None
        margin_ec = None
        r_mean = r_mic
        margin_mean = margin_mic
        sigma_i = SIGMA_FLOOR
        winners_agree = None

    p_win = phi(margin_mean * 100.0 / sigma_i)

    r_2024 = g24_row["reform_share_2024"] if g24_row else None
    swing = (r_mean - r_2024) * 100 if r_2024 is not None else None

    out_rows.append({
        "constituency": mic_row["name"],
        "projected_winner_mic": mic_row["winner"],
        "projected_winner_ec": ec_row["winner"] if ec_row else "",
        "winners_agree": ("" if winners_agree is None else
                          ("yes" if winners_agree else "no")),
        "reform_share_mean_pct": round(r_mean * 100, 2),
        "reform_share_mic_pct": round(r_mic * 100, 2),
        "reform_share_ec_pct": round(r_ec * 100, 2) if r_ec is not None else "",
        "reform_share_2024_pct": round(r_2024 * 100, 2) if r_2024 is not None else "",
        "reform_swing_pp": round(swing, 2) if swing is not None else "",
        "top_competitor": top_mic_party,
        "top_competitor_share_pct": round(top_mic_share * 100, 2),
        "margin_mean_pp": round(margin_mean * 100, 2),
        "margin_mic_pp": round(margin_mic * 100, 2),
        "margin_ec_pp": round(margin_ec * 100, 2) if margin_ec is not None else "",
        "sigma_pp": round(sigma_i, 2),
        "reform_win_probability": round(p_win, 3),
        "likelihood_category": categorise(p_win),
        "con_pct": round(mic_row["shares"]["Conservative"] * 100, 2),
        "lab_pct": round(mic_row["shares"]["Labour"] * 100, 2),
        "libdem_pct": round(mic_row["shares"]["Liberal Democrat"] * 100, 2),
        "green_pct": round(mic_row["shares"]["The Green Party"] * 100, 2),
        "snp_pct": round(mic_row["shares"]["Scottish National Party (SNP)"] * 100, 2),
        "plaid_pct": round(mic_row["shares"]["Plaid Cymru"] * 100, 2),
        "other_pct": round(mic_row["shares"]["Other"] * 100, 2),
        "winner_2024": g24_row["winner_2024"] if g24_row else "",
    })

out_rows.sort(key=lambda x: -x["reform_win_probability"])

fields = list(out_rows[0].keys())
with open(OUT, "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=fields)
    w.writeheader()
    w.writerows(out_rows)

n = len(out_rows)
mic_wins = sum(1 for r in out_rows if r["projected_winner_mic"] == "Reform UK")
ec_wins = sum(1 for r in out_rows if r["projected_winner_ec"] == "Reform")
both = sum(1 for r in out_rows if r["winners_agree"] == "yes" and r["projected_winner_mic"] == "Reform UK")
disputed = sum(1 for r in out_rows if r["winners_agree"] == "no")
safe = sum(1 for r in out_rows if r["likelihood_category"] == "Safe Reform")
likely = sum(1 for r in out_rows if r["likelihood_category"] == "Likely Reform")
tossup = sum(1 for r in out_rows if r["likelihood_category"] == "Toss-up")
expected = sum(r["reform_win_probability"] for r in out_rows)

print(f"\nRows written: {n}")
print(f"MiC projects Reform wins:     {mic_wins}")
print(f"EC projects Reform wins:      {ec_wins}")
print(f"Both agree Reform wins:       {both}")
print(f"Pollsters disagree on winner: {disputed} seats")
print(f"Expected Reform seats (ΣP):   {expected:.1f}")
print(f"Safe Reform:   {safe}")
print(f"Likely Reform: {likely}")
print(f"Toss-up:       {tossup}")
